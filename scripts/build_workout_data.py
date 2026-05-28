#!/usr/bin/env python3
"""
Builds the workout exercise bank.

Source of truth for v1: the EXERCISES list below (Python).
Outputs:
  - Resources/workout.xlsx              (authoring spreadsheet, rewritten)
  - PlankApp/Resources/exercises.json   (bundled into app)

Workflow after v1:
  - Edit Resources/workout.xlsx OR the EXERCISES list.
  - Re-run this script.
  - When `--from-xlsx` is passed, the Excel becomes the source and JSON is
    regenerated from it (use this once the spreadsheet is the easier place
    to edit).
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = REPO_ROOT / "Resources" / "workout.xlsx"
JSON_PATH = REPO_ROOT / "PlankApp" / "Resources" / "exercises.json"
SWIFT_PATH = REPO_ROOT / "PlankApp" / "Workout" / "ExerciseBankData.swift"


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------

# target_areas (multi):
#   abs | obliques | lowerBack | glutes | quads | hamstrings | hipFlexors |
#   calves | upperBody | fullBody
# type (single): cardio | strength | core | mobility | balance
# impact (single): low | med | high
# difficulty: 1..5
# symmetry: bilateral | alternating | unilateral
# default_side: left | right | None (only when symmetry == unilateral)
# pace: hold | rep — drives audio cueing (countdown/tempo lines)
# met: float (Compendium of Physical Activities reference)


# Body position taxonomy — drives the generator's block-based ordering
# (Pamela Reif / growingannanas convention). The block sequence per
# session is: standing → quadruped → plank → prone → sideLying → supine
# → seated. Each exercise gets exactly one position. Side-lying is
# always paired L/R within its block.
POSITIONS: dict[str, str] = {
    # === Posterior chain / back / upper-body anchors ===
    "plank_saw":              "plank",
    "alternating_superman":   "prone",
    "bird_dog":               "quadruped",
    "floor_dip":              "seated",
    "jumping_lunges":         "standing",
    "kneeling_shoulder_tap":  "quadruped",
    "mountain_climbers":      "plank",
    "back_extension_lying":   "prone",
    "superman_pull_up":       "prone",
    "superman_hold":          "prone",
    "w_raise":                "prone",
    "y_raise":                "prone",
    # === Abs / core ===
    "bent_knee_hip_raise":    "supine",
    "boat_flutters":          "seated",
    "boat_bicycle":           "seated",
    "cocoon_crunch":          "supine",
    "crunch_knee_raise":      "supine",
    "alt_knee_raise_crunch":  "supine",
    "crunch":                 "supine",
    "dead_bug":               "supine",
    "dead_bug_leg_lower":     "supine",
    "diagonal_plank":         "plank",
    "flutter_kicks":          "supine",
    "glute_bridge_march":     "supine",
    "bicycle_crunch":         "supine",
    "alt_leg_raise":          "supine",
    "leg_raise_hold":         "supine",
    "leg_raise":              "supine",
    "leg_raise_hip_lift":     "supine",
    "reverse_crunch":         "supine",
    "seated_knee_tuck":       "seated",
    "side_plank":             "sideLying",
    "side_crunch":            "sideLying",
    "side_crunch_hip_raise":  "sideLying",
    "sit_up":                 "supine",
    "standing_hip_abduction": "standing",
    "standing_side_bend":     "standing",
    "tabletop_bridge":        "supine",
    "tabletop_bridge_knee_lift": "supine",
    "tabletop_hold_knee_lift": "quadruped",
    "v_up":                   "supine",
    "vertical_leg_crunch":    "supine",
    "windshield_wipers":      "supine",
    # === Lower body / legs / glutes ===
    "reverse_to_forward_lunge": "standing",
    "air_squat":              "standing",
    "donkey_kick":            "quadruped",
    "donkey_kick_pulse":      "quadruped",
    "donkey_kickback":        "quadruped",
    "fire_hydrant":           "quadruped",
    "alt_forward_lunge":      "standing",
    "glute_bridge":           "supine",
    "good_morning":           "standing",
    "high_plank_leg_raise":   "plank",
    "kneeling_squat":         "quadruped",
    "forward_lunge":          "standing",
    "reverse_lunge":          "standing",
    "low_plank_leg_raise":    "plank",
    "narrow_squat":           "standing",
    "overhead_forward_lunge": "standing",
    "overhead_reverse_lunge": "standing",
    "side_lunge":             "standing",
    "side_lying_hip_abduction": "sideLying",
    "side_plank_hip_abduction": "sideLying",
    "side_split_squat":       "standing",
    "single_leg_glute_bridge": "supine",
    "single_leg_rdl":         "standing",
    "stiff_leg_deadlift":     "standing",
    "split_squat":            "standing",
    "squat":                  "standing",
    "squat_calf_raise":       "standing",
    "pulse_squat":            "standing",
    "calf_raise":             "standing",
    "single_leg_calf_raise":  "standing",
    "sumo_squat":             "standing",
    "wall_sit":               "standing",
    # === Cardio ===
    "skipping":               "standing",
    "bodyweight_swing":       "standing",
    "burpee":                 "standing",
    "burpee_no_jump":         "standing",
    "burpee_pushup":          "standing",
    "butt_kicks":             "standing",
    "crab_toe_touches":       "seated",
    "front_kicks":            "standing",
    "high_knees":             "standing",
    "jumping_jacks":          "standing",
    "modified_burpee":        "standing",
    "modified_jacks":         "standing",
    "plank_frog_jump":        "plank",
    "plank_jacks":            "plank",
    "boxing_punches":         "standing",
    "run_punch":              "standing",
    "run_in_place":           "standing",
    "plank_run":              "plank",
    "side_jacks":             "standing",
    "side_leg_jack":          "standing",
    "simple_jacks":           "standing",
    "jump_squat":             "standing",
    "squat_front_kick":       "standing",
    "sumo_jump_squat":        "standing",
    # === Mobility / stretching ===
    "baby_cobra":             "prone",
    "bow_pose":               "prone",
    "cat_cow":                "quadruped",
    "cat_stretch":            "quadruped",
    "childs_pose":            "quadruped",
    "cobra":                  "prone",
    "cow_stretch":            "quadruped",
    "shoulder_stretch":       "standing",
    "downward_dog":           "quadruped",
    "puppy_pose":             "quadruped",
    "forward_fold_bind":      "standing",
    "neck_stretch":           "standing",
    "sphinx":                 "prone",
    "side_tilt":              "standing",
    "backbend":               "standing",
    "upward_dog":             "prone",
    "reverse_tabletop":       "supine",
    "deep_hip_flexor":        "quadruped",
    "hamstring_stretch_standing": "standing",
    "hip_flexor_stretch":     "quadruped",
    "knees_to_chest":         "supine",
    "lizard_pose":            "quadruped",
    "lying_glute_stretch":    "supine",
    "lying_hamstring_stretch": "supine",
    "single_knee_to_chest":   "supine",
    "kneeling_quad_stretch":  "seated",
    "seated_forward_fold":    "seated",
    "standing_knee_hug":      "standing",
    "standing_quad_stretch":  "standing",
}

# Stillness-dominant exercises. Everything else is rep-based.
HOLD_IDS = {
    # core holds
    "side_plank", "superman_hold", "diagonal_plank", "wall_sit",
    "leg_raise_hold", "tabletop_hold_knee_lift",
    # all mobility/yoga
    "baby_cobra", "bow_pose", "cat_cow", "cat_stretch", "childs_pose", "cobra",
    "cow_stretch", "shoulder_stretch", "downward_dog", "puppy_pose",
    "forward_fold_bind", "neck_stretch", "sphinx", "side_tilt", "backbend",
    "upward_dog", "reverse_tabletop", "deep_hip_flexor",
    "hamstring_stretch_standing", "hip_flexor_stretch", "knees_to_chest",
    "lizard_pose", "lying_glute_stretch", "lying_hamstring_stretch",
    "single_knee_to_chest", "kneeling_quad_stretch", "seated_forward_fold",
    # balance holds (stretches with single-leg balance)
    "standing_knee_hug", "standing_quad_stretch",
}


@dataclass
class Exercise:
    id: str
    name: str
    lottie_id: int
    lottie_url: str
    target_areas: list[str]
    type: str
    impact: str
    difficulty: int
    met: float
    symmetry: str
    default_side: Optional[str] = None
    default_duration_sec: int = 30
    rest_after_sec: int = 15
    note: str = ""
    pace: str = "rep"            # set in validate() from HOLD_IDS
    position: str = ""           # set in validate() from POSITIONS
    lottie_file: str = ""        # set in validate() — filename stem under PlankApp/Resources/lottie


def E(
    id, name, lottie_id, lottie_url, areas, kind, impact, diff, met,
    sym, side=None, dur=30, rest=15, note="",
):
    return Exercise(
        id=id, name=name, lottie_id=lottie_id, lottie_url=lottie_url,
        target_areas=areas, type=kind, impact=impact, difficulty=diff,
        met=met, symmetry=sym, default_side=side,
        default_duration_sec=dur, rest_after_sec=rest, note=note,
    )


_BASE = "https://iconscout.com/lottie-animation/woman-doing-"


def L(slug, lid):
    return f"{_BASE}{slug}_{lid}"


# ---------------------------------------------------------------------------
# Exercise bank — 128 entries
# ---------------------------------------------------------------------------

EXERCISES: list[Exercise] = [

    # === Posterior chain / back / upper-body anchors ===
    E("plank_saw", "Plank Saw", 12236021,
      L("body-saw-exercise-for-for-shoulders-animation", 12236021),
      ["abs", "upperBody"], "core", "low", 3, 4.5, "bilateral"),
    E("alternating_superman", "Alternating Superman", 12236051,
      L("alternating-superman-exercise-for-back-animation", 12236051),
      ["lowerBack"], "strength", "low", 2, 3.5, "alternating"),
    E("bird_dog", "Bird Dog", 12236050,
      L("bird-dog-exercise-for-back-animation", 12236050),
      ["lowerBack", "abs"], "balance", "low", 2, 3.5, "alternating"),
    E("floor_dip", "Floor Dip", 12236008,
      L("floor-dip-exercise-for-arm-animation", 12236008),
      ["upperBody"], "strength", "low", 3, 5.0, "bilateral"),
    E("jumping_lunges", "Jumping Lunges", 12236015,
      L("jumping-lunges-cardio-exercise-animation", 12236015),
      ["quads", "glutes"], "cardio", "high", 4, 9.0, "alternating"),
    E("kneeling_shoulder_tap", "Kneeling Shoulder Tap", 12236025,
      L("kneeling-shoulder-tap-exercise-for-chest-animation", 12236025),
      ["abs", "upperBody"], "core", "low", 2, 4.5, "alternating"),
    E("mountain_climbers", "Mountain Climbers", 12236048,
      L("mountain-climber-exercise-for-abs-animation", 12236048),
      ["abs", "fullBody"], "cardio", "high", 3, 8.0, "alternating"),
    E("back_extension_lying", "Back Extension", 12236049,
      L("lying-back-extension-exercise-for-back-animation", 12236049),
      ["lowerBack"], "strength", "low", 2, 3.5, "bilateral"),
    E("superman_pull_up", "Superman Pull-Up", 12236054,
      L("superman-pull-up-exercise-for-back-animation", 12236054),
      ["lowerBack", "upperBody"], "strength", "low", 2, 4.0, "bilateral"),
    E("superman_hold", "Superman Hold", 12236055,
      L("superman-exercise-for-back-animation", 12236055),
      ["lowerBack"], "core", "low", 2, 3.5, "bilateral"),
    E("w_raise", "W Raise", 12236053,
      L("w-back-extension-exercise-for-back-animation", 12236053),
      ["lowerBack", "upperBody"], "strength", "low", 2, 3.5, "bilateral"),
    E("y_raise", "Y Raise", 12236052,
      L("y-back-extension-exercise-for-back-animation", 12236052),
      ["lowerBack", "upperBody"], "strength", "low", 2, 3.5, "bilateral"),

    # === Abs / core ===
    E("bent_knee_hip_raise", "Bent Knee Hip Raise", 12193330,
      L("bent-knee-hip-raise-exercise-for-abs-animation", 12193330),
      ["abs"], "core", "low", 2, 4.5, "bilateral"),
    E("boat_flutters", "Boat Hold + Flutter", 12193339,
      L("boat-hold-leg-flutters-exercise-for-abs-animation", 12193339),
      ["abs"], "core", "low", 4, 5.5, "bilateral"),
    E("boat_bicycle", "Boat Hold Bicycle", 12193331,
      L("boat-hold-bicycle-exercise-for-abs-animation", 12193331),
      ["abs", "obliques"], "core", "low", 4, 5.5, "alternating"),
    E("cocoon_crunch", "Cocoon Crunch", 12193332,
      L("cocoons-exercise-for-abs-animation", 12193332),
      ["abs"], "core", "low", 3, 5.0, "bilateral"),
    E("crunch_knee_raise", "Crunch + Knee Raise", 12193334,
      L("crunch-knee-raise-exercise-for-abs-animation", 12193334),
      ["abs"], "core", "low", 2, 4.5, "bilateral"),
    E("alt_knee_raise_crunch", "Alternating Knee Raise Crunch", 12193306,
      L("crunch-hold-knee-raise-alternating-exercise-for-abs-animation", 12193306),
      ["abs"], "core", "low", 3, 5.0, "alternating"),
    E("crunch", "Crunch", 12193333,
      L("crunch-exercise-for-abs-animation", 12193333),
      ["abs"], "core", "low", 1, 4.0, "bilateral"),
    E("dead_bug", "Dead Bug", 12193326,
      L("dead-bug-exercise-for-abs-animation", 12193326),
      ["abs"], "core", "low", 2, 4.0, "alternating"),
    E("dead_bug_leg_lower", "Dead Bug Leg Lower", 12193327,
      L("dead-bug-leg-lowering-exercise-for-abs-animation", 12193327),
      ["abs"], "core", "low", 3, 4.5, "alternating"),
    E("diagonal_plank", "Diagonal Plank", 12193319,
      L("diagonal-plank-exercise-for-abs-animation", 12193319),
      ["abs", "obliques"], "core", "low", 3, 5.0, "bilateral"),
    E("flutter_kicks", "Flutter Kicks", 12193338,
      L("flutter-kicks-exercise-for-abs-animation", 12193338),
      ["abs"], "core", "low", 2, 4.5, "alternating"),
    E("glute_bridge_march", "Glute Bridge March", 12193349,
      L("glute-bridge-march-exercise-for-legs-and-core-animation", 12193349),
      ["glutes", "abs"], "core", "low", 2, 4.5, "alternating"),
    E("bicycle_crunch", "Bicycle Crunch", 12193328,
      L("lying-bicycle-exercise-for-abs-animation", 12193328),
      ["abs", "obliques"], "core", "low", 2, 5.0, "alternating"),
    E("alt_leg_raise", "Alternating Leg Raise", 12193336,
      L("lying-leg-raise-alternating-exercise-for-abs-animation", 12193336),
      ["abs"], "core", "low", 2, 4.5, "alternating"),
    E("leg_raise_hold", "Leg Raise + Hold", 12193337,
      L("lying-leg-raise-and-hold-exercise-for-abs-animation", 12193337),
      ["abs"], "core", "low", 3, 5.0, "bilateral"),
    E("leg_raise", "Leg Raise", 12193344,
      L("lying-leg-raise-exercise-for-abs-animation", 12193344),
      ["abs"], "core", "low", 2, 4.5, "bilateral"),
    E("leg_raise_hip_lift", "Leg Raise + Hip Lift", 12193348,
      L("lying-leg-raise-with-hip-lift-exercise-for-abs-animation", 12193348),
      ["abs"], "core", "low", 4, 5.5, "bilateral"),
    E("reverse_crunch", "Reverse Crunch", 12193342,
      L("reverse-crunch-exercise-for-abs-animation", 12193342),
      ["abs"], "core", "low", 2, 4.5, "bilateral"),
    E("seated_knee_tuck", "Seated Knee Tuck", 12193311,
      L("seated-knee-tuck-exercise-for-abs-animation", 12193311),
      ["abs"], "core", "low", 2, 4.5, "bilateral"),
    E("side_plank", "Side Plank", 12193318,
      L("side-bridge-exercise-for-abs-animation", 12193318),
      ["obliques"], "core", "low", 3, 4.0, "unilateral", side="right",
      note="Mirror animation horizontally to render the left side."),
    E("side_crunch", "Side Crunch", 12193317,
      L("side-crunches-exercise-for-abs-animation", 12193317),
      ["obliques"], "core", "low", 2, 4.5, "unilateral", side="right",
      note="Mirror animation horizontally to render the left side."),
    E("side_crunch_hip_raise", "Side Crunch + Hip Raise", 12193316,
      L("side-crunches-hip-raise-exercise-for-abs-animation", 12193316),
      ["obliques"], "core", "low", 3, 5.0, "unilateral", side="right",
      note="Mirror animation horizontally to render the left side."),
    E("sit_up", "Sit-Up", 12193315,
      L("sit-up-exercise-for-abs-animation", 12193315),
      ["abs"], "core", "low", 2, 5.0, "bilateral"),
    E("standing_hip_abduction", "Standing Hip Abduction", 12193347,
      L("standing-hip-abduction-exercise-for-legs-animation", 12193347),
      ["glutes"], "strength", "low", 2, 4.0, "alternating"),
    E("standing_side_bend", "Standing Side Bend", 12193324,
      L("standing-side-bend-exercise-for-abs-animation", 12193324),
      ["obliques"], "mobility", "low", 1, 3.0, "unilateral", side="right",
      note="Mirror animation horizontally to render the left side."),
    E("tabletop_bridge", "Tabletop Bridge", 12193353,
      L("tabletop-bridge-exercise-for-legs-animation", 12193353),
      ["glutes", "abs"], "strength", "low", 2, 4.0, "bilateral"),
    E("tabletop_bridge_knee_lift", "Tabletop Bridge + Knee Lift", 12193352,
      L("tabletop-bridge-with-bent-knee-lift-exercise-for-legs-animation", 12193352),
      ["glutes", "abs"], "strength", "low", 3, 4.5, "alternating"),
    E("tabletop_hold_knee_lift", "Tabletop Hold + Knee Lift", 12193351,
      L("tabletop-hold-with-bent-knee-lift-exercise-for-legs-animation", 12193351),
      ["glutes", "abs"], "balance", "low", 3, 4.5, "alternating"),
    E("v_up", "V-Up", 12193322,
      L("v-up-exercise-for-abs-animation", 12193322),
      ["abs"], "core", "low", 4, 5.5, "bilateral"),
    E("vertical_leg_crunch", "Vertical Leg Crunch", 12193321,
      L("vertical-leg-crunch-exercise-for-abs-animation", 12193321),
      ["abs"], "core", "low", 3, 5.0, "bilateral"),
    E("windshield_wipers", "Windshield Wipers", 12193329,
      L("wipers-exercise-for-abs-animation", 12193329),
      ["abs", "obliques"], "core", "low", 4, 5.5, "alternating"),

    # === Lower body / legs / glutes ===
    E("reverse_to_forward_lunge", "Reverse-to-Forward Lunge", 12193231,
      L("back-to-forward-lunge-exercise-for-legs-animation", 12193231),
      ["quads", "glutes"], "strength", "low", 3, 6.0, "unilateral", side="right",
      note="Mirror animation horizontally to render the left side."),
    E("air_squat", "Air Squat", 12193252,
      L("air-squat-exercise-for-legs-animation", 12193252),
      ["quads", "glutes"], "strength", "low", 1, 5.0, "bilateral"),
    E("donkey_kick", "Donkey Kick", 12193232,
      L("donkey-kick-exercise-for-legs-animation", 12193232),
      ["glutes"], "strength", "low", 1, 4.0, "unilateral", side="left",
      note="Mirror animation horizontally to render the right side."),
    E("donkey_kick_pulse", "Donkey Kick Pulse", 12193220,
      L("donkey-kick-pulse-exercise-for-legs-animation", 12193220),
      ["glutes"], "strength", "low", 2, 4.5, "unilateral", side="left",
      note="Mirror animation horizontally to render the right side."),
    E("donkey_kickback", "Donkey Kickback", 12193233,
      L("donkey-kickback-exercise-for-legs-animation", 12193233),
      ["glutes"], "strength", "low", 2, 4.5, "unilateral", side="left",
      note="Mirror animation horizontally to render the right side."),
    E("fire_hydrant", "Fire Hydrant", 12193219,
      L("fire-hydrant-exercise-for-legs-animation", 12193219),
      ["glutes"], "strength", "low", 2, 4.5, "unilateral", side="left",
      note="Mirror animation horizontally to render the right side."),
    E("alt_forward_lunge", "Alternating Forward Lunge", 12193234,
      L("forward-lunges-alternating-exercise-for-legs-animation", 12193234),
      ["quads", "glutes"], "strength", "low", 2, 5.5, "alternating"),
    E("glute_bridge", "Glute Bridge", 12193209,
      L("glute-bridge-exercise-for-legs-animation", 12193209),
      ["glutes"], "strength", "low", 1, 4.0, "bilateral"),
    E("good_morning", "Good Morning", 12193218,
      L("good-morning-exercise-for-legs-animation", 12193218),
      ["hamstrings", "lowerBack"], "strength", "low", 2, 4.0, "bilateral"),
    E("high_plank_leg_raise", "High Plank Leg Raise", 12193217,
      L("high-plank-leg-raise-exercise-for-legs-animation", 12193217),
      ["glutes", "abs"], "core", "low", 3, 5.5, "alternating"),
    E("kneeling_squat", "Kneeling Squat", 12193207,
      L("kneeling-squat-exercise-for-legs-animation", 12193207),
      ["glutes", "quads"], "strength", "low", 2, 4.5, "bilateral"),
    E("forward_lunge", "Forward Lunge", 12193235,
      L("left-leg-forward-lunge-exercise-for-legs-animation", 12193235),
      ["quads", "glutes"], "strength", "low", 2, 5.0, "unilateral", side="left",
      note="Mirror animation horizontally to render the right side."),
    E("reverse_lunge", "Reverse Lunge", 12193238,
      L("left-leg-reverse-lunge-exercise-for-legs-animation", 12193238),
      ["quads", "glutes"], "strength", "low", 2, 5.0, "unilateral", side="left",
      note="Mirror animation horizontally to render the right side."),
    E("low_plank_leg_raise", "Low Plank Leg Raise", 12193216,
      L("low-plank-leg-raise-exercise-for-legs-animation", 12193216),
      ["glutes", "abs"], "core", "low", 4, 5.5, "alternating"),
    E("narrow_squat", "Narrow Squat", 12193250,
      L("narrow-squat-exercise-for-legs-animation", 12193250),
      ["quads"], "strength", "low", 2, 5.0, "bilateral"),
    E("overhead_forward_lunge", "Overhead Forward Lunge", 12193227,
      L("overhead-forward-lunges-exercise-for-legs-animation", 12193227),
      ["quads", "glutes", "upperBody"], "strength", "low", 3, 6.0, "alternating"),
    E("overhead_reverse_lunge", "Overhead Reverse Lunge", 12193236,
      L("overhead-reverse-lunges-exercise-for-legs-animation", 12193236),
      ["quads", "glutes", "upperBody"], "strength", "low", 3, 6.0, "alternating"),
    E("side_lunge", "Side Lunge", 12193243,
      L("side-lunge-exercise-for-legs-animation", 12193243),
      ["glutes", "quads"], "strength", "low", 2, 5.5, "unilateral", side="left",
      note="Mirror animation horizontally to render the right side."),
    E("side_lying_hip_abduction", "Side-Lying Hip Abduction", 12193225,
      L("side-lying-hip-abduction-exercise-for-legs-animation", 12193225),
      ["glutes"], "strength", "low", 1, 3.5, "unilateral", side="left",
      note="Mirror animation horizontally to render the right side."),
    E("side_plank_hip_abduction", "Side Plank Hip Abduction", 12193224,
      L("side-plank-hip-abduction-exercise-for-legs-animation", 12193224),
      ["obliques", "glutes"], "core", "low", 4, 5.0, "unilateral", side="left",
      note="Mirror animation horizontally to render the right side."),
    E("side_split_squat", "Side Split Squat", 12193242,
      L("side-split-squat-exercise-for-legs-animation", 12193242),
      ["quads", "glutes"], "strength", "low", 3, 5.5, "unilateral", side="left",
      note="Mirror animation horizontally to render the right side."),
    E("single_leg_glute_bridge", "Single-Leg Glute Bridge", 12193208,
      L("single-leg-glute-bridge-exercise-for-legs-animation", 12193208),
      ["glutes"], "strength", "low", 2, 4.5, "unilateral", side="left",
      note="Mirror animation horizontally to render the right side."),
    E("single_leg_rdl", "Single-Leg RDL", 12193223,
      L("single-leg-romanian-deadlift-exercise-for-legs-animation", 12193223),
      ["hamstrings", "glutes"], "balance", "low", 4, 5.0, "unilateral", side="left",
      note="Mirror animation horizontally to render the right side."),
    E("stiff_leg_deadlift", "Stiff-Leg Deadlift", 12193222,
      L("single-stiff-leg-deadlift-exercise-for-legs-animation", 12193222),
      ["hamstrings", "glutes"], "balance", "low", 4, 5.0, "unilateral", side="left",
      note="Mirror animation horizontally to render the right side."),
    E("split_squat", "Split Squat", 12193246,
      L("split-squat-exercise-for-legs-animation", 12193246),
      ["quads", "glutes"], "strength", "low", 3, 5.5, "unilateral", side="right",
      note="Mirror animation horizontally to render the left side."),
    E("squat", "Squat", 12193255,
      L("squat-exercise-for-legs-animation", 12193255),
      ["quads", "glutes"], "strength", "low", 2, 5.5, "bilateral"),
    E("squat_calf_raise", "Squat to Calf Raise", 12193254,
      L("squat-to-calf-raise-exercise-for-legs-animation", 12193254),
      ["quads", "glutes", "calves"], "strength", "low", 2, 5.5, "bilateral"),
    E("pulse_squat", "Pulse Squat", 12193221,
      L("squat-with-pulse-exercise-for-legs-animation", 12193221),
      ["quads", "glutes"], "strength", "low", 3, 6.0, "bilateral"),
    E("calf_raise", "Calf Raise", 12193229,
      L("standing-calf-raise-exercise-for-legs-animation", 12193229),
      ["calves"], "strength", "low", 1, 3.0, "bilateral"),
    E("single_leg_calf_raise", "Single-Leg Calf Raise", 12193213,
      L("standing-single-leg-calf-raise-exercise-for-legs-animation", 12193213),
      ["calves"], "balance", "low", 2, 3.5, "unilateral", side="right",
      note="Mirror animation horizontally to render the left side."),
    E("sumo_squat", "Sumo Squat", 12193253,
      L("sumo-squat-exercise-for-legs-animation", 12193253),
      ["quads", "glutes"], "strength", "low", 2, 5.5, "bilateral"),
    E("wall_sit", "Wall Sit", 12193241,
      L("wall-sit-exercise-for-legs-animation", 12193241),
      ["quads"], "core", "low", 3, 5.0, "bilateral"),

    # === Cardio ===
    E("skipping", "Skipping", 12349335,
      L("bodyweight-skipping-cardio-exercise-animation", 12349335),
      ["calves", "fullBody"], "cardio", "med", 2, 7.5, "bilateral"),
    E("bodyweight_swing", "Bodyweight Swing", 12349324,
      L("bodyweight-swing-cardio-exercise-animation", 12349324),
      ["fullBody"], "cardio", "low", 2, 6.0, "bilateral"),
    E("burpee", "Burpee", 12349325,
      L("burpee-cardio-exercise-animation", 12349325),
      ["fullBody"], "cardio", "high", 4, 9.0, "bilateral"),
    E("burpee_no_jump", "Burpee (No Jump)", 12349326,
      L("burpee-no-jump-cardio-exercise-animation", 12349326),
      ["fullBody"], "cardio", "med", 3, 7.5, "bilateral"),
    E("burpee_pushup", "Burpee + Push-Up", 12349327,
      L("burpee-with-push-up-cardio-exercise-animation", 12349327),
      ["fullBody", "upperBody"], "cardio", "high", 5, 10.0, "bilateral"),
    E("butt_kicks", "Butt Kicks", 12349313,
      L("butt-kicks-cardio-exercise-animation", 12349313),
      ["hamstrings"], "cardio", "med", 2, 7.0, "alternating"),
    E("crab_toe_touches", "Crab Toe Touches", 12349314,
      L("crab-toe-touches-cardio-exercise-animation", 12349314),
      ["abs", "upperBody"], "cardio", "low", 3, 6.0, "alternating"),
    E("front_kicks", "Front Kicks", 12349315,
      L("front-kicks-cardio-exercise-animation", 12349315),
      ["quads", "abs"], "cardio", "med", 2, 6.0, "alternating"),
    E("high_knees", "High Knees", 12349316,
      L("high-knee-taps-cardio-exercise-animation", 12349316),
      ["quads", "abs"], "cardio", "med", 2, 7.0, "alternating"),
    E("jumping_jacks", "Jumping Jacks", 12349337,
      L("jumping-jack-cardio-exercise-animation", 12349337),
      ["fullBody"], "cardio", "high", 2, 8.0, "bilateral"),
    E("modified_burpee", "Modified Burpee", 12349328,
      L("modified-burpee-cardio-exercise-animation", 12349328),
      ["fullBody"], "cardio", "low", 2, 6.0, "bilateral"),
    E("modified_jacks", "Modified Jumping Jacks", 12349338,
      L("modified-jumping-jack-cardio-exercise-animation", 12349338),
      ["fullBody"], "cardio", "low", 1, 5.0, "bilateral"),
    E("plank_frog_jump", "Plank Frog Jump", 12349329,
      L("plank-frog-jump-cardio-exercise-animation", 12349329),
      ["abs", "fullBody"], "cardio", "med", 4, 8.0, "bilateral"),
    E("plank_jacks", "Plank Jacks", 12349317,
      L("plank-jack-cardio-exercise-animation", 12349317),
      ["abs", "fullBody"], "cardio", "med", 3, 7.0, "bilateral"),
    E("boxing_punches", "Boxing Punches", 12349320,
      L("punches-cardio-exercise-animation", 12349320),
      ["upperBody"], "cardio", "low", 2, 6.0, "alternating"),
    E("run_punch", "Run + Punch", 12349318,
      L("running-in-place-and-punches-cardio-exercise-animation", 12349318),
      ["fullBody"], "cardio", "med", 2, 7.0, "alternating"),
    E("run_in_place", "Run in Place", 12349319,
      L("running-in-place-cardio-exercise-animation", 12349319),
      ["quads"], "cardio", "med", 2, 7.0, "alternating"),
    E("plank_run", "Plank Run", 12349322,
      L("running-plank-cardio-exercise-animation", 12349322),
      ["abs", "fullBody"], "cardio", "med", 3, 7.5, "alternating"),
    E("side_jacks", "Side Jacks", 12349332,
      L("side-jack-cardio-exercise-animation", 12349332),
      ["fullBody"], "cardio", "med", 2, 7.0, "alternating"),
    E("side_leg_jack", "Side Leg Jack", 12349333,
      L("side-leg-raise-jack-cardio-exercise-animation", 12349333),
      ["glutes"], "cardio", "med", 3, 7.0, "alternating"),
    E("simple_jacks", "Simple Jumping Jacks", 12349334,
      L("simplified-jumping-jack-cardio-exercise-animation", 12349334),
      ["fullBody"], "cardio", "low", 1, 5.0, "bilateral"),
    E("jump_squat", "Jump Squat", 12349310,
      L("squat-jump-cardio-exercise-animation", 12349310),
      ["quads", "glutes"], "cardio", "high", 4, 8.5, "bilateral"),
    E("squat_front_kick", "Squat + Front Kick", 12349311,
      L("squat-with-front-kicks-cardio-exercise-animation", 12349311),
      ["quads", "abs"], "cardio", "med", 3, 7.0, "alternating"),
    E("sumo_jump_squat", "Sumo Jump Squat", 12349312,
      L("sumo-squat-jump-cardio-exercise-animation", 12349312),
      ["quads", "glutes"], "cardio", "high", 4, 8.5, "bilateral"),

    # === Mobility / stretching (yoga + classic) ===
    E("baby_cobra", "Baby Cobra", 12682305,
      L("baby-cobra-pose-stretching-animation", 12682305),
      ["lowerBack"], "mobility", "low", 1, 2.3, "bilateral", dur=20),
    E("bow_pose", "Bow Pose", 12682302,
      L("bow-pose-stretching-animation", 12682302),
      ["lowerBack", "fullBody"], "mobility", "low", 3, 2.5, "bilateral", dur=20),
    E("cat_cow", "Cat-Cow", 12682301,
      L("cat-cow-stretching-animation", 12682301),
      ["lowerBack", "abs"], "mobility", "low", 1, 2.3, "alternating", dur=30),
    E("cat_stretch", "Cat Stretch", 12682300,
      L("cat-stretching-animation", 12682300),
      ["lowerBack"], "mobility", "low", 1, 2.3, "bilateral", dur=20),
    E("childs_pose", "Child's Pose", 12682309,
      L("childs-pose-back-stretching-animation", 12682309),
      ["lowerBack"], "mobility", "low", 1, 2.3, "bilateral", dur=30),
    E("cobra", "Cobra", 12682308,
      L("cobra-pose-stretching-animation", 12682308),
      ["lowerBack"], "mobility", "low", 1, 2.5, "bilateral", dur=20),
    E("cow_stretch", "Cow Stretch", 12682307,
      L("cow-stretching-animation", 12682307),
      ["lowerBack"], "mobility", "low", 1, 2.3, "bilateral", dur=20),
    E("shoulder_stretch", "Shoulder Stretch", 12682290,
      L("deep-shoulders-stretching-animation", 12682290),
      ["upperBody"], "mobility", "low", 1, 2.3, "bilateral", dur=20),
    E("downward_dog", "Downward Dog", 12682292,
      L("downward-facing-dog-stretching-animation", 12682292),
      ["fullBody", "upperBody"], "mobility", "low", 2, 2.8, "bilateral", dur=30),
    E("puppy_pose", "Puppy Pose", 12682293,
      L("extended-puppy-pose-stretching-animation", 12682293),
      ["upperBody"], "mobility", "low", 1, 2.3, "bilateral", dur=20),
    E("forward_fold_bind", "Forward Fold + Bind", 12682294,
      L("forward-bending-hands-behind-stretching-animation", 12682294),
      ["hamstrings"], "mobility", "low", 2, 2.5, "bilateral", dur=20),
    E("neck_stretch", "Neck Stretch", 12682284,
      L("neck-and-trapezius-stretching-animation", 12682284),
      ["upperBody"], "mobility", "low", 1, 2.3, "unilateral", side="right", dur=20,
      note="Mirror animation horizontally to render the left side."),
    E("sphinx", "Sphinx", 12682288,
      L("sphinx-pose-stretching-animation", 12682288),
      ["lowerBack"], "mobility", "low", 1, 2.3, "bilateral", dur=20,
      note="Excel notes only-left; sphinx is symmetric — bilateral."),
    E("side_tilt", "Side Tilt", 12682289,
      L("side-tilt-stretching-animation", 12682289),
      ["obliques"], "mobility", "low", 1, 2.3, "alternating", dur=20),
    E("backbend", "Backbend", 12682287,
      L("spinal-extension-backward-bending-stretching-animation", 12682287),
      ["lowerBack"], "mobility", "low", 2, 2.5, "bilateral", dur=20),
    E("upward_dog", "Upward Dog", 12682276,
      L("upward-facing-dog-stretching-animation", 12682276),
      ["lowerBack"], "mobility", "low", 2, 2.5, "bilateral", dur=20),
    E("reverse_tabletop", "Reverse Tabletop", 12682306,
      L("upward-table-pose-stretching-animation", 12682306),
      ["fullBody"], "mobility", "low", 3, 3.0, "bilateral", dur=20),
    E("deep_hip_flexor", "Deep Hip Flexor Stretch", 12443462,
      L("advanced-hip-flexor-stretching-animation", 12443462),
      ["hipFlexors"], "mobility", "low", 3, 2.5, "unilateral", side="right", dur=25,
      note="Mirror animation horizontally to render the left side."),
    E("hamstring_stretch_standing", "Standing Hamstring Stretch", 12443465,
      L("hamstring-stretching-animation", 12443465),
      ["hamstrings"], "mobility", "low", 1, 2.3, "unilateral", side="left", dur=25,
      note="Mirror animation horizontally to render the right side."),
    E("hip_flexor_stretch", "Hip Flexor Stretch", 12443466,
      L("hip-flexor-stretching-animation", 12443466),
      ["hipFlexors"], "mobility", "low", 2, 2.3, "unilateral", side="left", dur=25,
      note="Mirror animation horizontally to render the right side."),
    E("knees_to_chest", "Knees to Chest", 12443453,
      L("knees-to-chest-glute-stretching-animation", 12443453),
      ["lowerBack", "glutes"], "mobility", "low", 1, 2.3, "bilateral", dur=20),
    E("lizard_pose", "Lizard Pose", 12443455,
      L("lizard-pose-stretching-animation", 12443455),
      ["hipFlexors"], "mobility", "low", 3, 2.5, "alternating", dur=20),
    E("lying_glute_stretch", "Lying Glute Stretch", 12443457,
      L("lying-glute-stretching-animation", 12443457),
      ["glutes"], "mobility", "low", 1, 2.3, "unilateral", side="left", dur=25,
      note="Mirror animation horizontally to render the right side."),
    E("lying_hamstring_stretch", "Lying Hamstring Stretch", 12443458,
      L("lying-hamstring-stretching-animation", 12443458),
      ["hamstrings"], "mobility", "low", 1, 2.3, "unilateral", side="right", dur=25,
      note="Mirror animation horizontally to render the left side."),
    E("single_knee_to_chest", "Single Knee to Chest", 12443451,
      L("single-knee-to-chest-glute-stretching-animation", 12443451),
      ["lowerBack", "glutes"], "mobility", "low", 1, 2.3, "unilateral", side="right", dur=25,
      note="Mirror animation horizontally to render the left side."),
    E("kneeling_quad_stretch", "Kneeling Quad Stretch", 12443447,
      L("sitting-kneeling-quadriceps-stretching-animation", 12443447),
      ["quads"], "mobility", "low", 2, 2.3, "alternating", dur=25),
    E("seated_forward_fold", "Seated Forward Fold", 12443448,
      L("spinal-flexion-forward-bending-stretching-animation", 12443448),
      ["lowerBack", "hamstrings"], "mobility", "low", 1, 2.3, "bilateral", dur=25),
    E("standing_knee_hug", "Standing Knee Hug", 12443460,
      L("standing-knee-to-chest-glute-stretching-animation", 12443460),
      ["lowerBack", "glutes"], "balance", "low", 2, 2.5, "unilateral", side="right", dur=20,
      note="Mirror animation horizontally to render the left side."),
    E("standing_quad_stretch", "Standing Quad Stretch", 12443461,
      L("standing-quadricep-stretching-animation", 12443461),
      ["quads"], "balance", "low", 2, 2.5, "unilateral", side="right", dur=20,
      note="Mirror animation horizontally to render the left side."),
]


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

VALID_AREAS = {
    "abs", "obliques", "lowerBack", "glutes", "quads", "hamstrings",
    "hipFlexors", "calves", "upperBody", "fullBody",
}
VALID_TYPES = {"cardio", "strength", "core", "mobility", "balance"}
VALID_IMPACTS = {"low", "med", "high"}
VALID_SYMMETRY = {"bilateral", "alternating", "unilateral"}
VALID_SIDES = {"left", "right"}
VALID_PACE = {"hold", "rep"}
VALID_POSITIONS = {"standing", "quadruped", "plank", "prone", "sideLying", "supine", "seated"}


def validate():
    seen_ids: set[str] = set()
    seen_lottie: set[int] = set()
    for ex in EXERCISES:
        if ex.id in seen_ids:
            raise ValueError(f"Duplicate id: {ex.id}")
        seen_ids.add(ex.id)
        if ex.lottie_id in seen_lottie:
            raise ValueError(f"Duplicate lottie_id: {ex.lottie_id} ({ex.id})")
        seen_lottie.add(ex.lottie_id)
        for area in ex.target_areas:
            if area not in VALID_AREAS:
                raise ValueError(f"{ex.id}: bad area {area}")
        if ex.type not in VALID_TYPES:
            raise ValueError(f"{ex.id}: bad type {ex.type}")
        if ex.impact not in VALID_IMPACTS:
            raise ValueError(f"{ex.id}: bad impact {ex.impact}")
        if not (1 <= ex.difficulty <= 5):
            raise ValueError(f"{ex.id}: bad difficulty {ex.difficulty}")
        if ex.symmetry not in VALID_SYMMETRY:
            raise ValueError(f"{ex.id}: bad symmetry {ex.symmetry}")
        if ex.symmetry == "unilateral":
            if ex.default_side not in VALID_SIDES:
                raise ValueError(f"{ex.id}: unilateral needs default_side")
        else:
            if ex.default_side is not None:
                raise ValueError(f"{ex.id}: non-unilateral cannot have default_side")
        ex.pace = "hold" if ex.id in HOLD_IDS else "rep"
        if ex.pace not in VALID_PACE:
            raise ValueError(f"{ex.id}: bad pace {ex.pace}")

        # Position drives the generator's block-based ordering. Every
        # exercise must be tagged.
        if ex.id not in POSITIONS:
            raise ValueError(f"{ex.id}: missing position tag in POSITIONS dict")
        ex.position = POSITIONS[ex.id]
        if ex.position not in VALID_POSITIONS:
            raise ValueError(f"{ex.id}: bad position {ex.position}")

        # Derive lottie filename from the URL slug (matches the manual download
        # naming: woman-doing-{slug}-animation_{id} → woman-doing-{slug})
        import re as _re
        m = _re.search(r"lottie-animation/(woman-doing-[a-z0-9-]+?)-animation_\d+$", ex.lottie_url)
        if not m:
            raise ValueError(f"{ex.id}: cannot derive lottie filename from {ex.lottie_url}")
        ex.lottie_file = m.group(1)

    # Sanity: HOLD_IDS shouldn't reference non-existent exercises.
    unknown_holds = HOLD_IDS - seen_ids
    if unknown_holds:
        raise ValueError(f"HOLD_IDS references unknown exercises: {sorted(unknown_holds)}")

    holds = sum(1 for e in EXERCISES if e.pace == "hold")
    print(f"OK — {len(EXERCISES)} exercises validated ({holds} hold, {len(EXERCISES)-holds} rep).")


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------

XLSX_HEADERS = [
    "id", "name", "lottie_id", "lottie_url", "lottie_file",
    "target_areas", "type", "impact", "difficulty",
    "met", "symmetry", "default_side", "pace", "position",
    "default_duration_sec", "rest_after_sec", "note",
]


def write_xlsx(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "exercises"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4F4F")
    for col_idx, h in enumerate(XLSX_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, ex in enumerate(EXERCISES, start=2):
        row = [
            ex.id, ex.name, ex.lottie_id, ex.lottie_url, ex.lottie_file,
            ", ".join(ex.target_areas), ex.type, ex.impact, ex.difficulty,
            ex.met, ex.symmetry, ex.default_side or "", ex.pace, ex.position,
            ex.default_duration_sec, ex.rest_after_sec, ex.note,
        ]
        for col_idx, v in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    widths = [22, 32, 12, 70, 50, 26, 10, 8, 10, 6, 14, 12, 6, 12, 10, 10, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {path} ({len(EXERCISES)} rows).")


def write_json(path: Path):
    payload = {
        "version": 1,
        "schema": {
            "target_areas": sorted(VALID_AREAS),
            "type": sorted(VALID_TYPES),
            "impact": sorted(VALID_IMPACTS),
            "symmetry": sorted(VALID_SYMMETRY),
            "pace": sorted(VALID_PACE),
            "difficulty_scale": [1, 5],
        },
        "exercises": [asdict(ex) for ex in EXERCISES],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2) + "\n")
    print(f"Wrote {path} ({len(EXERCISES)} exercises).")


def _swift_str(s: str) -> str:
    return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'


def _swift_areas(areas: list[str]) -> str:
    return "[" + ", ".join(f".{a}" for a in areas) + "]"


def write_swift(path: Path):
    lines: list[str] = [
        "// AUTO-GENERATED by Scripts/build_workout_data.py — DO NOT EDIT.",
        "// Source of truth: Resources/workout.xlsx",
        "//",
        f"// {len(EXERCISES)} exercises.",
        "",
        "import Foundation",
        "",
        "enum ExerciseBankData {",
        "    static let all: [Exercise] = [",
    ]
    for ex in EXERCISES:
        side = f".{ex.default_side}" if ex.default_side else "nil"
        note = _swift_str(ex.note)
        lines.append("        Exercise(")
        lines.append(f"            id: {_swift_str(ex.id)},")
        lines.append(f"            name: {_swift_str(ex.name)},")
        lines.append(f"            lottieId: {ex.lottie_id},")
        lines.append(f"            lottieURL: {_swift_str(ex.lottie_url)},")
        lines.append(f"            targetAreas: {_swift_areas(ex.target_areas)},")
        lines.append(f"            type: .{ex.type},")
        lines.append(f"            impact: .{ex.impact},")
        lines.append(f"            difficulty: {ex.difficulty},")
        lines.append(f"            met: {ex.met},")
        lines.append(f"            symmetry: .{ex.symmetry},")
        lines.append(f"            defaultSide: {side},")
        lines.append(f"            pace: .{ex.pace},")
        lines.append(f"            position: .{ex.position},")
        lines.append(f"            lottieFile: {_swift_str(ex.lottie_file)},")
        lines.append(f"            defaultDurationSec: {ex.default_duration_sec},")
        lines.append(f"            restAfterSec: {ex.rest_after_sec},")
        lines.append(f"            note: {note}")
        lines.append("        ),")
    lines.append("    ]")
    lines.append("}")
    lines.append("")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines))
    print(f"Wrote {path} ({len(EXERCISES)} exercises).")


# ---------------------------------------------------------------------------
# Reader (xlsx → list[Exercise]) — used when Excel is the source of truth
# ---------------------------------------------------------------------------

def read_from_xlsx(path: Path) -> list[Exercise]:
    wb = openpyxl.load_workbook(path)
    ws = wb["exercises"] if "exercises" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    if headers != XLSX_HEADERS:
        raise ValueError(f"Unexpected xlsx headers: {headers}")
    out: list[Exercise] = []
    for r in rows[1:]:
        if r[0] in (None, ""):
            continue
        d = dict(zip(headers, r))
        out.append(Exercise(
            id=d["id"],
            name=d["name"],
            lottie_id=int(d["lottie_id"]),
            lottie_url=d["lottie_url"],
            target_areas=[s.strip() for s in str(d["target_areas"]).split(",") if s.strip()],
            type=d["type"],
            impact=d["impact"],
            difficulty=int(d["difficulty"]),
            met=float(d["met"]),
            symmetry=d["symmetry"],
            default_side=(d["default_side"] or None) or None,
            default_duration_sec=int(d["default_duration_sec"]),
            rest_after_sec=int(d["rest_after_sec"]),
            note=d["note"] or "",
            pace=d.get("pace") or "rep",
            lottie_file=d.get("lottie_file") or "",
        ))
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--from-xlsx", action="store_true",
                        help="Read EXERCISES from xlsx instead of the Python source.")
    args = parser.parse_args()

    global EXERCISES  # noqa: PLW0603
    if args.from_xlsx:
        EXERCISES = read_from_xlsx(XLSX_PATH)
        print(f"Loaded {len(EXERCISES)} exercises from {XLSX_PATH}")

    validate()
    write_xlsx(XLSX_PATH)
    write_json(JSON_PATH)
    write_swift(SWIFT_PATH)


if __name__ == "__main__":
    main()
